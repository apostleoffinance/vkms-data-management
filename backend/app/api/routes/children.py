import io
import uuid
from datetime import date

from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook

from app.core.deps import AdminUser, DbSession, VerifiedUser
from app.models.child import Child, Gender
from app.models.class_model import Class
from app.schemas.child import (
    ChildCreate,
    ChildDetailResponse,
    ChildResponse,
    ChildSearchResult,
    ChildUpdate,
)
from app.services.audit import log_audit
from app.services.bulk_import_service import (
    ImportColumnMap,
    build_column_map,
    looks_like_header_row,
    parse_import_row,
)
from app.services.child_service import (
    find_class_by_name,
    generate_child_code,
    generate_qr_code,
    get_child_detail,
    get_or_create_parent,
    is_empty_placeholder,
    search_children,
)
from app.services.pickup_service import create_contact, ensure_primary_contact_from_parent

router = APIRouter()


def _child_response(child: Child, parent_linked: bool = False) -> ChildResponse:
    return ChildResponse(
        id=str(child.id),
        child_code=child.child_code,
        first_name=child.first_name,
        last_name=child.last_name,
        gender=child.gender.value,
        date_of_birth=child.date_of_birth,
        parent_id=str(child.parent_id),
        class_id=str(child.class_id),
        medical_notes=child.medical_notes,
        registration_date=child.registration_date,
        is_active=child.is_active,
        qr_code_data=child.qr_code_data,
        created_at=child.created_at.isoformat(),
        updated_at=child.updated_at.isoformat(),
        parent_linked=parent_linked,
    )


@router.get("/search", response_model=list[ChildSearchResult])
def search(
    q: str = Query(default="", max_length=100),
    missing_photos: bool = Query(default=False),
    include_inactive: bool = Query(default=False),
    limit: int = Query(default=50, ge=1, le=100),
    db: DbSession = ...,
    current_user: VerifiedUser = ...,
) -> list[ChildSearchResult]:
    if missing_photos and current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins can filter by missing pickup photos",
        )
    if include_inactive and current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only admins can include inactive children",
        )
    results = search_children(
        db,
        q,
        limit=limit,
        include_inactive=include_inactive,
        missing_photos_only=missing_photos,
    )
    return [ChildSearchResult(**r) for r in results]


@router.get("/qr/{child_code}", response_model=ChildSearchResult)
def search_by_qr(child_code: str, db: DbSession, current_user: VerifiedUser) -> ChildSearchResult:
    child = db.query(Child).filter(Child.child_code == child_code).first()
    if not child:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Child not found")
    results = search_children(db, child.child_code, limit=1)
    if not results:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Child not found")
    return ChildSearchResult(**results[0])


@router.get("/{child_id}", response_model=ChildDetailResponse)
def get_child(child_id: uuid.UUID, db: DbSession, current_user: VerifiedUser) -> ChildDetailResponse:
    detail = get_child_detail(db, child_id)
    if not detail:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Child not found")
    return ChildDetailResponse(**detail)


@router.post("", response_model=ChildResponse, status_code=status.HTTP_201_CREATED)
def register_child(body: ChildCreate, db: DbSession, admin: AdminUser) -> ChildResponse:
    class_ = db.query(Class).filter(Class.id == uuid.UUID(body.class_id)).first()
    if not class_:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class not found")

    parent_uuid = None
    if body.parent_id:
        try:
            parent_uuid = uuid.UUID(body.parent_id)
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid parent_id")

    try:
        parent, parent_created = get_or_create_parent(
            db,
            first_name=body.parent_first_name,
            last_name=body.parent_last_name,
            phone=body.parent_phone,
            alternative_phone=body.parent_alternative_phone,
            email=body.parent_email,
            address=body.parent_address,
            parent_id=parent_uuid,
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc

    child_code = generate_child_code(db)
    child_id = uuid.uuid4()
    qr_data = generate_qr_code(child_code, child_id)

    try:
        gender = Gender(body.gender)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid gender")

    child = Child(
        id=child_id,
        child_code=child_code,
        first_name=body.first_name,
        last_name=body.last_name,
        gender=gender,
        date_of_birth=body.date_of_birth,
        parent_id=parent.id,
        class_id=class_.id,
        medical_notes=body.medical_notes,
        registration_date=date.today(),
        qr_code_data=qr_data,
    )
    db.add(child)
    db.flush()

    if body.authorized_pickups:
        for index, pickup in enumerate(body.authorized_pickups):
            photo_data = None
            photo_content_type = None
            if pickup.photo_base64:
                from app.services.photo_service import decode_photo_base64

                try:
                    photo_data, photo_content_type = decode_photo_base64(pickup.photo_base64)
                except ValueError as exc:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
            create_contact(
                db,
                child_id=child.id,
                first_name=pickup.first_name,
                last_name=pickup.last_name,
                phone=pickup.phone,
                relationship=pickup.relationship,
                is_primary=pickup.is_primary or index == 0,
                photo_data=photo_data,
                photo_content_type=photo_content_type,
            )
    else:
        ensure_primary_contact_from_parent(db, child)

    db.commit()
    db.refresh(child)
    log_audit(
        db,
        "create",
        "child",
        user_id=admin.id,
        resource_id=str(child.id),
        details={"parent_linked": not parent_created, "parent_id": str(parent.id)},
    )
    return _child_response(child, parent_linked=not parent_created)


@router.put("/{child_id}", response_model=ChildResponse)
def update_child(
    child_id: uuid.UUID, body: ChildUpdate, db: DbSession, admin: AdminUser
) -> ChildResponse:
    child = db.query(Child).filter(Child.id == child_id).first()
    if not child:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Child not found")

    if body.first_name is not None:
        child.first_name = body.first_name
    if body.last_name is not None:
        child.last_name = body.last_name
    if body.gender is not None:
        child.gender = Gender(body.gender)
    if body.date_of_birth is not None:
        child.date_of_birth = body.date_of_birth
    if body.class_id is not None:
        class_ = db.query(Class).filter(Class.id == uuid.UUID(body.class_id)).first()
        if not class_:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Class not found")
        child.class_id = class_.id
    if body.medical_notes is not None:
        child.medical_notes = body.medical_notes
    if body.is_active is not None:
        child.is_active = body.is_active

    db.commit()
    db.refresh(child)
    log_audit(db, "update", "child", user_id=admin.id, resource_id=str(child.id))
    return _child_response(child)


@router.post("/bulk-import", status_code=status.HTTP_201_CREATED)
def bulk_import(
    db: DbSession,
    admin: AdminUser,
    file: UploadFile = File(...),
) -> dict:
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel file required")

    content = file.file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {"created": 0, "errors": ["Spreadsheet is empty"]}

    headers = tuple(all_rows[0])
    col_map: ImportColumnMap = (
        build_column_map(headers) if looks_like_header_row(headers) else ImportColumnMap()
    )
    data_rows = all_rows[1:]
    created = 0
    errors = []

    for i, row in enumerate(data_rows, start=2):
        if not row:
            continue
        first_col = col_map.child_first_name if col_map.child_first_name is not None else 0
        if is_empty_placeholder(row[first_col] if len(row) > first_col else None):
            continue
        try:
            parsed = parse_import_row(tuple(row), col_map)

            class_ = find_class_by_name(db, parsed.class_name)
            if not class_:
                label = parsed.class_name or "default class"
                errors.append(f"Row {i}: Class '{label}' not found")
                continue

            existing_child = (
                db.query(Child)
                .filter(
                    Child.first_name.ilike(parsed.child_first_name),
                    Child.last_name.ilike(parsed.child_last_name),
                    Child.date_of_birth == parsed.date_of_birth,
                )
                .first()
            )
            if existing_child:
                errors.append(f"Row {i}: Child already registered (skipped)")
                continue

            try:
                parent, _ = get_or_create_parent(
                    db,
                    first_name=parsed.parent.first_name,
                    last_name=parsed.parent.last_name,
                    phone=parsed.parent.phone,
                    alternative_phone=parsed.parent.alternative_phone,
                    email=parsed.parent.email,
                    address=parsed.parent.address,
                )
            except ValueError as exc:
                errors.append(f"Row {i}: {exc}")
                continue

            child_code = generate_child_code(db)
            child_id = uuid.uuid4()
            qr_data = generate_qr_code(child_code, child_id)

            try:
                gender = Gender(parsed.gender.lower())
            except ValueError:
                errors.append(f"Row {i}: Invalid gender '{parsed.gender}'")
                continue

            child = Child(
                id=child_id,
                child_code=child_code,
                first_name=parsed.child_first_name,
                last_name=parsed.child_last_name,
                gender=gender,
                date_of_birth=parsed.date_of_birth,
                parent_id=parent.id,
                class_id=class_.id,
                medical_notes=parsed.medical_notes,
                registration_date=date.today(),
                qr_code_data=qr_data,
            )
            db.add(child)
            db.commit()
            ensure_primary_contact_from_parent(db, child)
            db.commit()
            created += 1
        except ValueError as exc:
            errors.append(f"Row {i}: {exc}")
        except Exception as e:
            db.rollback()
            errors.append(f"Row {i}: {str(e)}")

    log_audit(
        db, "bulk_import", "child", user_id=admin.id, details={"created": created, "errors": errors}
    )
    return {"created": created, "errors": errors}
